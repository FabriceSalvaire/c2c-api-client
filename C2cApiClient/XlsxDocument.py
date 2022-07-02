####################################################################################################
#
# C2cApiClient - A Python client for the camptocamp.org API
# Copyright (C) 2017 Salvaire Fabrice
# Contact: http://www.fabrice-salvaire.fr
# SPDX-License-Identifier: AGPL-3.0-only
#
####################################################################################################

__all__ = ['XlsxDocument']

####################################################################################################

from openpyxl import Workbook

####################################################################################################

class XlsxDocument:

    ##############################################

    def __init__(self, titles):
        self._wb = Workbook()
        self._ws1 = self._wb.active
        self._ws1.title = "Routes"
        self._row = 0
        self._set_column_titles(titles)

    ##############################################

    @property
    def current_row(self):
        return self._row
    
    ##############################################

    def save(self, path: str):
        self._wb.save(filename=path)

    ##############################################

    def add_row(self, values, row=None):
        if row is None:
            # pre-increment for current row
            self._row += 1
            row = self._row
        else:
            self._row = row
        col = 1
        for value in values:
            self._ws1.cell(column=col, row=row, value=value)
            col += 1

    ##############################################

    def _set_column_titles(self, titles):
        self._columns = titles
        self.add_row(titles, row=1)

    ##############################################

    def column_for_title(self, title):
        return self._columns.index(title) +1

    ##############################################

    def cell_for_title(self, title):
        row = self._row
        col = self.column_for_title(title)
        print(row, col)
        return self._ws1.cell(row=row, column=col)
